import calendar
import openpyxl
from datetime import timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def main():
    print("Programa fet per Albert Constans Badosa.\nNo canviar el nom al fitxer Excel.\nNo afegir cap columna a horesUF ni a Festes.\nQuan acabi d'executar el programa autoadjusteu la mida de les columnes i de les files a la pestanya Programació.\nLa pestanya programació es reiniciarà cada vegada que s'executi el programa.\nSi les dades d'input no estan bé el programa pot fallar i/o no acabar sense avisar que ha tingut un problema. Verifiqueu que els resultats siguin coherents.")
    
    font = Font(b=True)
    try:
        wb = openpyxl.load_workbook("programacio.xlsx")
    except:
        print("El fitxer programacio.xlsx no existeix.")
        return 1
    wspro = wb["hores UF"]
    wsfes = wb["Festes"]
    inici = wsfes["B1"].value.date()
    final = wsfes["C1"].value.date()
    
    calendari = calendar.Calendar()
    
    #importar festes
    festes = []
    i = 2
    while(True):    
        if(wsfes["B"+str(i)].value == None ):
            break
        if(wsfes["C"+str(i)].value == None):
            festes.append(wsfes["B"+str(i)].value.date())
        else:
            date = wsfes["B"+str(i)].value
            while(wsfes["C"+str(i)].value >= date):
                festes.append(date.date())
                date  += timedelta(days=1)
        i += 1
        
    #importar dates
    assignatures = []
    horesAssignatures = []
    horesFetesAssignatures = []
    assignaturesPrimer = []
    assignaturesUF = []
    dataFinal = []
    i = 2
    while(True):
        if(wspro["A"+str(i)].value == None):
            break
        if(wspro["B"+str(i)].value == None and wspro["A"+str(i)].value is not None):
            assignatures.append(wspro["A"+str(i)].value)
            temp = []
            temp.append(int(wspro["F"+str(i)].value))
            temp.append(int(wspro["G"+str(i)].value))
            temp.append(int(wspro["H"+str(i)].value))
            temp.append(int(wspro["I"+str(i)].value))
            temp.append(int(wspro["J"+str(i)].value))
            horesAssignatures.append(temp)
            horesFetesAssignatures.append(0)
            if(wspro["C"+str(i)].value == 1):
                assignaturesPrimer.append(True)
            else:
                assignaturesPrimer.append(False)
            temp = []
            temp2 = []
            j = 1
            while(True):
                if(wspro["B"+str(i+j)].value is not None and wspro["A"+str(i+j)].value is not None):  
                    temp.append([wspro["A"+str(i+j)].value, int(wspro["B"+str(i+j)].value)])
                    temp2.append(0)
                    if(j is not 1):
                        temp[j-1][1] += temp[j-2][1]
                else:
                    break
                j += 1
            assignaturesUF.append(temp)
            dataFinal.append(temp2)
        i += 1  
    
    if 'Programacio' in wb.sheetnames:
        wb.remove(wb["Programacio"])
        
    wsout = wb.create_sheet("Programacio")
    wsout.cell(1, 1, "Curs").font = font
    wsout.cell(1, 2, str(inici.year)+"-"+str(final.year)).font = font
    try:
        s = input("Nom del cicle: ")
    except:
        s = "GRAU MITJÀ EN INSTAL·LACIONS ELÈCTRIQUES I AUTOMÀTIQUES"
    wsout.cell(3, 1, s).font = font
    wsout.merge_cells("A3:D3")
    wsout.cell(5, 1, "Organització dels continguts en mòduls").font = font
    wsout.merge_cells("A5:B5")
    wsout.cell(6, 1, "TOTAL HORES CICLE").font = font
    a = 0
    for assig in assignaturesUF:
        a += assig[-1][1]
    wsout.cell(6, 2, a)
    wsout.cell(9, 1, "Mòduls 1er Curs").font = font
    a = 0
    b = 0
    c = sum(assignaturesPrimer)
    wsout.cell(17+c+1, 1, "Distribució dels crèdits al llarg del curs").font = font
    wsout.cell(17+c+3, 1, "Mòduls 2on Curs").font = font
    wsout.cell(28+len(assignaturesPrimer)+1, 1, "Distribució dels crèdits al llarg del crus").font = font
    wsout.cell(16, 1, "Nom assignatura").font = font
    wsout.cell(16, 2, "Hores setmana").font = font
    wsout.cell(16, 3, "Hores assignatura").font = font
    wsout.cell(33, 1, "Nom assignatura").font = font
    wsout.cell(33, 2, "Hores setmana").font = font
    wsout.cell(33, 3, "Hores assignatura").font = font
    wsout.cell(11, 2, "Setmanes:").font = font
    wsout.cell(12, 2, "h/setmanes:").font = font
    wsout.cell(13, 2, "h totals:").font = font
    suma = 0
    for i in range(len(assignaturesPrimer)):
        if(assignaturesPrimer[i]):
            suma += sum(horesAssignatures[i])    
    wsout.cell(12, 3, suma)
    suma = 0
    for i in range(len(assignaturesPrimer)):
        if(not assignaturesPrimer[i]):
            suma += sum(horesAssignatures[i])         
    wsout.cell(28, 2, "Setmanes:").font = font
    wsout.cell(29, 2, "h/setmanes:").font = font
    wsout.cell(29, 3, suma)
    wsout.cell(30, 2, "h totals:").font = font
    for i in range(len(assignatures)):
        if(assignaturesPrimer[i]):
            wsout.cell(17+a, 1, assignatures[i])
            wsout.cell(17+a, 2, sum(horesAssignatures[i]))
            wsout.cell(17+a, 3, assignaturesUF[i][-1][1])
            a += 1
        else:
            wsout.cell(28+c+b, 1, assignatures[i])
            wsout.cell(28+c+b, 2, sum(horesAssignatures[i]))
            wsout.cell(28+c+b, 3, assignaturesUF[i][-1][1])
            b += 1
    
    j = 47
    
    pos = []
    done = 0
    thin = Side(border_style="thin", color="000000")
    border = Border(right = thin)
    for i in range(len(assignatures)):
        wsout.cell(j, 1, assignatures[i]).font = font
        pos.append(j+2)
        l = sum(horesAssignatures[i])
        if(l == 0):
            l = 1
        for k in range(l):
            wsout.cell(j+3+k, 4, k+1).border = border
        j += sum(horesAssignatures[i]) + 4
    
    ali = Alignment(textRotation=90)
    setmana = -1
    border = Border(bottom = thin)
    borderleft = Border(left = thin)
    borderright = Border(right = thin)
    bordertop = Border(top = thin)
    bordermax = Border(right = thin, bottom = thin)
    fill = PatternFill(fill_type="solid", start_color='FFFFFF')
    wsout.row_dimensions[9].border = border
    wsout.row_dimensions[26].border = border
    wsout["A9"].border = border
    wsout["A26"].border = border
    
    #Iterar dies
    for v in range (2):
        for month in range(1, 13):        
            for days in calendari.itermonthdates(inici.year+v, month):
                if(days.month != month): 
                    continue
                if(days.weekday() == 0):
                    done = 0
                    for p in range(len(horesFetesAssignatures)):
                        if(dataFinal[p][-1] == 0 and setmana >= 0):
                            l = sum(horesAssignatures[p])
                            if(l == 0):
                                l = 1
                            wsout.cell(pos[p]+l+1, 5+setmana).border = bordertop
                if(days < inici or days in festes or days.weekday() > 4):
                    continue
                if(days > final):
                    break
                for i in range(len(horesFetesAssignatures)):
                    #Afegir hores cada dia
                    horesFetesAssignatures[i] += horesAssignatures[i][days.weekday()]
                    
                    if(done == 0):
                        done = 1
                        setmana += 1
                        for p in range(len(horesFetesAssignatures)):
                            if(dataFinal[p][-1] == 0):
                                wsout.cell(pos[p], 5+setmana, days-timedelta(days=days.weekday())).border = border
                                wsout.cell(pos[p], 5+setmana).alignment = ali 
                                l = sum(horesAssignatures[p])
                                if(l == 0):
                                    l = 1
                                for h in range(l):
                                    wsout.cell(pos[p]+h+1, 5+setmana).fill = fill
                    
                    #Mirar si una UF s'ha acabat
                    for j in range(len(assignaturesUF[i])):
                        if(dataFinal[i][j] == 0 and assignaturesUF[i][j][1] <= horesFetesAssignatures[i]):
                            dataFinal[i][j] = days 
                            m = sum(horesAssignatures[i][0:days.weekday()+1])-1+(assignaturesUF[i][j][1]-horesFetesAssignatures[i])
                            if (days.weekday() == 0):
                                l = sum(horesAssignatures[i])
                                if(l == 0):
                                    l = 1
                                for k in range(l):
                                    if(k < m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = borderright 
                                    if(k == m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = bordermax
                                    if(k > m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = borderleft
                            else: 
                                l = sum(horesAssignatures[i])
                                if(l == 0):
                                    l = 1
                                for k in range(l):
                                    if(k < m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = borderright 
                                    if(k == m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = bordermax
                                    if(k > m):
                                        wsout.cell(pos[i]+1+k, 5+setmana).border = borderleft
                            
                #print str(days) + " - " + str(days.weekday())
                            
    #border = Border(left = thin)
    #for i in range(len(pos)):
    #    if(dataFinal[i][-1] == 0):
    #        l = sum(horesAssignatures[i])
    #        if(l == 0):
    #            l = 1
    #        for k in range(l):
    #            wsout.cell(pos[i]+1+k, 6+setmana).border = border
    
    wsout.cell(28, 3, setmana)
    wsout.cell(30, 3, setmana*wsout.cell(29, 3).value)
    wsout.cell(11, 3, setmana)
    wsout.cell(13, 3, setmana*wsout.cell(12, 3).value)                        
                  
    wb.save("programacio.xlsx")
    return 0
    
if __name__ == "__main__":
    if(main() == 0):
        print("Finalitzat sense cap problema.")
    input("Pitxeu Enter per sortir")